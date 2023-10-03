import datetime
import locale
import requests
import json
import math
import openpyxl
import sys
import datetime


locale.setlocale(locale.LC_TIME, "it_IT")  # italian

soglia1 = 0
soglia2 = 0
soglia3 = 0
soglia4 = 0
soglia5 = 0
soglia6 = 0
soglia7 = 0
soglia8 = 0
soglia9 = 0
soglia10 = 0
soglia13 = 0
soglia14 = 0
soglia15 = 0
soglia16 = 0
soglia17 = 0
soglia18 = 0
soglia19 = 0
soglia20 = 0
soglia21 = 0
soglia22 = 0
soglia23 = 0
soglia24 = 0
soglia25 = 0
soglia26 = 0
soglia27 = 0
soglia28 = 0
soglia29 = 0
soglia30 = 0
soglia31 = 0
soglia32 = 0
soglia33 = 0
soglia34 = 0
soglia35 = 0
soglia36 = 0

TOKEN_ID = '157136-1N4kvJex2PbVqB'
SPORT_ID = '1'

now = datetime.datetime.now()
print("Current date and time : ")
print(now.strftime("%Y-%m-%d %H:%M:%S"))


class Game:
    def __init__(self, home, away, time, result):
        self.home = home
        self.away = away
        self.time = time
        self.result = result

    def printGame(self):
        print('home: ' + self.home + ' away: ' + self.away + ' final: ' + self.result + ' time: ' + self.time)

    def __str__(self):
        return self.home+'-'+self.away+' '+self.result + ' '+ self.time

    def print(self):
        return self.home+'-'+self.away+' '+self.result + ' '+ self.time


def getOddMatch(idEvent):
    response = requests.get(
        'https://api.b365api.com/v2/event/odds/summary?token=' + TOKEN_ID + '&event_id=' + idEvent)
    j = json.loads(response.text)
    string_res = ""
    if len(j['results']) > 0:
        # todo se è null odds_update, prendere STARTS
        try:
            level = len(j['results']['BetFair']['odds_update'])
            if level == 0:
                level = 1
            if len(j['results']['BetFair']['odds']['start']) > 0:
                where = "start"
            else:
                where = "end"

            odds_X = float(j['results']['BetFair']['odds'][where]['1_' + str(level)]['draw_od'])
            odds_1 = float(j['results']['BetFair']['odds'][where]['1_' + str(level)]['home_od'])
            odds_2 = float(j['results']['BetFair']['odds'][where]['1_' + str(level)]['away_od'])
            string_res = "1: " + str(odds_1) + " X: " + str(odds_X) + " 2: " + str(odds_2)
        except Exception as e:
            print("ERRORE: ")
            print(e)
            string_res= "No BETFAIR ODDS"
    return string_res


def getNumberOfDayPlayed(id_league):
    response = requests.get('https://api.b365api.com/v3/league/table?token=' + TOKEN_ID + '&league_id='+id_league)
    league_json = json.loads(response.text)
    current_round = 0
    try:
        if len(league_json['results'][0]['overall']['tables']) > 1:
            current_round = league_json['results'][0]['overall']['tables'][0]['currentround']
    except Exception as e:
        print(e)

    return int(current_round)


def getPointTeam(id_league, id_team):
    response = requests.get('https://api.b365api.com/v3/league/table?token=' + TOKEN_ID + '&league_id='+id_league)
    league_json = json.loads(response.text)
    current_points = 0
    current_position = 0
    try:
        if len(league_json['results'][0]['overall']['tables']) > 1:
            table = league_json['results'][0]['overall']['tables'][0]['rows']
            for team in table:
                if team['team']['id'] == id_team:
                    current_points = team['points']
                    current_position = team['pos']
    except Exception as e:
        print(e)
        
    return current_points, current_position


def getEndedDayEvents(data="", campionato=""):
    now = datetime.datetime.now()
    now_str = now.strftime("%Y%m%d")
    if data != "":
        now_str = data

    print("DATA DELLE PARTITE ANALIZZATE: " + str(now_str))

    # tomorrow = now + datetime.timedelta(days=1)
    # print("DOMANI: ", tomorrow.strftime("%Y%m%d"))

    response = requests.get(
        'https://api.b365api.com/v3/events/ended?sport_id=' + SPORT_ID + '&token=' + TOKEN_ID + '&page=1&skip_esports=SI&day=' + now_str + '&cc='+campionato)
    print('https://api.b365api.com/v3/events/ended?sport_id=' + SPORT_ID + '&token=' + TOKEN_ID + '&page=1&skip_esports=SI&day=' + now_str + '&cc='+campionato)
    ended_json = json.loads(response.text)
    number_of_page = math.ceil(ended_json['pager']['total'] / ended_json['pager']['per_page'])

    for page in range(2, number_of_page + 1):
        r = requests.get(
            'https://api.b365api.com/v3/events/ended?sport_id=' + SPORT_ID + '&token=' + TOKEN_ID + '&page=' + str(
                page) + '&skip_esports=SI&day=' + now_str)
        j = json.loads(r.text)
        ended_json['results'] = ended_json['results'] + j['results']

    return ended_json


def getDayEvents(data="", campionato=""):
    now = datetime.datetime.now()
    now_str = now.strftime("%Y%m%d")
    if data != "":
        now_str = data

    print("DATA DELLE PARTITE ANALIZZATE: " + str(now_str))

    # tomorrow = now + datetime.timedelta(days=1)
    # print("DOMANI: ", tomorrow.strftime("%Y%m%d"))

    response = requests.get(
        'https://api.b365api.com/v3/events/upcoming?sport_id=' + SPORT_ID + '&token=' + TOKEN_ID + '&page=1&skip_esports=SI&day=' + now_str + '&cc='+campionato)

    upcoming_json = json.loads(response.text)
    number_of_page = math.ceil(upcoming_json['pager']['total'] / upcoming_json['pager']['per_page'])

    for page in range(2, number_of_page + 1):
        r = requests.get(
            'https://api.b365api.com/v3/events/upcoming?sport_id=' + SPORT_ID + '&token=' + TOKEN_ID + '&page=' + str(
                page) + '&skip_esports=SI&day=' + now_str)
        j = json.loads(r.text)
        upcoming_json['results'] = upcoming_json['results'] + j['results']

    return upcoming_json


def getHistoryForDayEvent(idEvent, home_team, away_team):
    response = requests.get('https://api.b365api.com/v1/event/history?token=' + TOKEN_ID + '&event_id=' + idEvent)
    j = json.loads(response.text)

    numero_precedenti = 0
    try:
        numero_precedenti = len(j['results']['home'])
    except Exception as e:
        print(e)

    favorita, favoritaNome, statoFormaUnder, res_string = statoForma(j['results']['home'], j['results']['away'],
                                                                     j['results']['h2h'], home_team, away_team)
    precedenti_string = h2h(j['results']['h2h'], home_team, away_team, favorita, favoritaNome)
    casa, vittorieCasa, sconfitteCasa, overCasa, underCasa = home_res(j['results']['home'], home_team, favorita)
    trasferta, vittorieOspite, sconfitteOspite, underOspite, overOspite = ris_away(j['results']['away'],
                                                                                                   away_team, favorita)

    return favorita, favoritaNome, statoFormaUnder, res_string, precedenti_string, vittorieCasa, sconfitteCasa, overCasa, underCasa, vittorieOspite, sconfitteOspite, underOspite, overOspite, numero_precedenti


def h2h(h2h_json, home_team, away_team, favorita, favoritaNome):
    precedentiOver25 = 0
    precedentiUnder25 = 0
    precedentiGoal = 0
    precedentiNoGoal = 0
    esito1 = 0
    esito2 = 0
    vittoriaF = 0

    totalePrec = 0

    home_score = 0
    away_score = 0

    for i in range(0, len(h2h_json)):
        if h2h_json[i]['home']['name'] == home_team:
            home_score += int(str(h2h_json[i]['ss']).split("-")[0])
            away_score += int(str(h2h_json[i]['ss']).split("-")[1])
        else:
            home_score += int(str(h2h_json[i]['ss']).split("-")[1])
            away_score += int(str(h2h_json[i]['ss']).split("-")[0])

        ris = home_score + away_score

        if home_score > 0 and away_score > 0:
            precedentiGoal = precedentiGoal + 1
            totalePrec = totalePrec + 1
        else:
            precedentiNoGoal = precedentiNoGoal + 1
            totalePrec = totalePrec + 1

        if (ris > 2):
            precedentiOver25 = precedentiOver25 + 1
        else:
            precedentiUnder25 = precedentiUnder25 + 1

        if home_score > away_score:
            esito1 = esito1 + 1
        if home_score < away_score:
            esito2 = esito2 + 1

        if home_team == favoritaNome and home_score > away_score:
            vittoriaF = vittoriaF + 1
        if away_team == favoritaNome and home_score < away_score:
            vittoriaF = vittoriaF + 1

    # Calcolo soglia
    try:
        precedentiOver25 = precedentiOver25 / totalePrec
        precedentiUnder25 = precedentiUnder25 / totalePrec
        precedentiGoal = precedentiGoal / totalePrec
        precedentiNoGoal = precedentiNoGoal / totalePrec
        esito1 = esito1 / totalePrec
        esito2 = esito2 / totalePrec
    except Exception as e:
        print("ERRORE precedenti division")
        print(e)

    res_precedenti_string = ""
    if precedentiOver25 > soglia5:
        # print('Segno Over')
        res_precedenti_string += "Over  \t"
    if precedentiUnder25 > soglia6:
        # print('Segno Under')
        res_precedenti_string += "Under \t"
    if precedentiGoal > soglia7:
        # print('Segno Goal')
        res_precedenti_string += "Goal  \t"
    if precedentiNoGoal > soglia8:
        # print('Segno NoGoal')
        res_precedenti_string += "NoGoal \t"
    if esito1 > soglia9:
        # print('Segno 1')
        res_precedenti_string += "1 \t"
    if esito2 > soglia10:
        # print('Segno 2')
        res_precedenti_string += "2 \t"

    if favorita == 1:
        if vittoriaF > soglia13:
            # print('Segno 1')
            res_precedenti_string += "1 \t"
    if favorita == 2:
        if vittoriaF > soglia14:
            # print('Segno 2')
            res_precedenti_string += "2 \t"
    else:
        # print('Nessuna Favorita Calcolata')
        res_precedenti_string += "Nessuna Favorita Calcolata \t"

    # print("############# RISULTATI PRECEDENTI ##########################")
    # print(res_precedenti_string)
    # print("#######################################")

    return res_precedenti_string


def statoForma(homeJs, awayJs, h2hJs, homeTeam, awayTeam):
    home_forma = 0
    away_forma = 0

    all_game = []
    for fonte in [homeJs, awayJs, h2hJs]:
        for js in fonte:
            g = Game(js['home']['name'], js['away']['name'], js['time'], js['ss'])
            all_game.append(g)

    all_game.sort(key=lambda x: x.time, reverse=True)
    all_game = all_game[:10]

    for j in all_game:
        if j.home == homeTeam:
            home_goal = j.result.split('-')[0]
            away_goal = j.result.split('-')[1]
            if home_goal > away_goal:
                home_forma += 3
            elif home_goal == away_goal:
                home_forma += 1
        elif j.away == homeTeam:
            home_goal = j.result.split('-')[0]
            away_goal = j.result.split('-')[1]
            if home_goal < away_goal:
                home_forma += 3
            elif home_goal == away_goal:
                home_forma += 1

        if j.home == awayTeam:
            home_goal = j.result.split('-')[0]
            away_goal = j.result.split('-')[1]
            if home_goal > away_goal:
                away_forma += 3
            elif home_goal == away_goal:
                away_forma += 1
        elif j.away == homeTeam:
            home_goal = j.result.split('-')[0]
            away_goal = j.result.split('-')[1]
            if home_goal < away_goal:
                away_forma += 3
            elif home_goal == away_goal:
                away_forma += 1


    # for j in homeJs:
    #     if j['home']['name'] == homeTeam:
    #         home_goal = j["ss"].split('-')[0]
    #         away_goal = j["ss"].split('-')[1]
    #         if home_goal > away_goal:
    #             home_forma += 3
    #         elif home_goal == away_goal:
    #             home_forma += 1
    #     elif j['away']['name'] == homeTeam:
    #         home_goal = j["ss"].split('-')[0]
    #         away_goal = j["ss"].split('-')[1]
    #         if home_goal < away_goal:
    #             home_forma += 3
    #         elif home_goal == away_goal:
    #             home_forma += 1
    #
    # for x in awayJs:
    #     if x['home']['name'] == awayTeam:
    #         home_goal = x["ss"].split('-')[0]
    #         away_goal = x["ss"].split('-')[1]
    #         if home_goal > away_goal:
    #             away_forma += 3
    #         elif home_goal == away_goal:
    #             away_forma += 1
    #     elif x['away']['name'] == homeTeam:
    #         home_goal = x["ss"].split('-')[0]
    #         away_goal = x["ss"].split('-')[1]
    #         if home_goal < away_goal:
    #             away_forma += 3
    #         elif home_goal == away_goal:
    #             away_forma += 1

    favorita = 0
    favoritaNome = 'nessuna Favorita'
    res_string = ""
    statoFormaUnder = 0

    if (home_forma - away_forma) > soglia3:
        res_string += "1/Over \t"
    elif (away_forma - home_forma) > soglia4:
        res_string += "2/Over \t"
    else:
        statoFormaUnder = 1
        res_string += "Under  \t"

    if home_forma > away_forma:  # Se forma1 > forma2 allora favorita = 1 squadra1
        favorita = 1
        favoritaNome = homeTeam
    elif home_forma < away_forma:
        favorita = 2
        favoritaNome = awayTeam

    return favorita, favoritaNome, statoFormaUnder, res_string


def home_res(home_j, home_team, favorita):
    vittorieCasa = 0  # Calcolo vittorie casa
    sconfitteCasa = 0  # Calcolo sconfitte casa
    pareggiCasa = 0  # Calcolo pareggi Casa
    overCasa = 0  # Calcolo Over Casa
    underCasa = 0  # Calcolo Under Casa

    partite = []

    home_game_sorted = []
    for js in home_j:
        g = Game(js['home']['name'], js['away']['name'], js['time'], js['ss'])
        home_game_sorted.append(g)

    home_game_sorted.sort(key=lambda x: x.time, reverse=True)
    home_game_sorted = home_game_sorted[:5]

    for match in home_game_sorted:
        home_score = 0
        other_score = 0
        partite.append({"data": match.time, 'squadre': match.home + '-' + match.away,
                        'risultato': match.result})
        if match.home == home_team:
            home_score = str(match.result).split("-")[0]
            other_score = str(match.result).split("-")[1]
        elif match.away == home_team:
            home_score = str(match.result).split("-")[1]
            other_score = str(match.result).split("-")[0]

        # partite.append({"data": match['time'], 'squadre': match['home']['name'] + '-' + match['away']['name'],
        #                 'risultato': match['ss']})
        # if match['home']['name'] == home_team:
        #     home_score = str(match['ss']).split("-")[0]
        #     other_score = str(match['ss']).split("-")[1]
        # elif match['away']['name'] == home_team:
        #     home_score = str(match['ss']).split("-")[1]
        #     other_score = str(match['ss']).split("-")[0]

        ris = int(home_score) + int(other_score)

        # Calcolo vittorie Casa, sconfitte casa ed over Casa ed underCasa
        if home_score > other_score:
            vittorieCasa = vittorieCasa + 1
        elif home_score < other_score:
            sconfitteCasa = sconfitteCasa + 1
        else:
            pareggiCasa = pareggiCasa + 1

        if (ris > 2):
            overCasa = overCasa + 1
        else:
            underCasa = underCasa + 1

    # Inserimento if se favorita = 1
    if favorita == 1:
        if vittorieCasa == 5:
            print('Casa 5 vittorie')
        if sconfitteCasa == 0:
            print('Casa non perde da 5')

    # Inserimento if se favorita = 2
    if favorita == 2:
        if sconfitteCasa == 5:
            print('Casa 5 sconfitte')
        if vittorieCasa == 0:
            print('Casa non vince da 5')

    return partite, vittorieCasa, sconfitteCasa, overCasa, underCasa


def ris_away(away_j, away_team, favorita):
    vittorieOspite = 0  # Calcolo vittorie ospite
    sconfitteOspite = 0  # Calcolo sconfitte ospite
    pareggiOspite = 0  # Calcolo pareggi ospite
    overOspite = 0  # Calcolo over ospite
    underOspite = 0  # Calcolo under ospite

    partite = []

    away_game_sorted = []
    for js in away_j:
        g = Game(js['home']['name'], js['away']['name'], js['time'], js['ss'])
        away_game_sorted.append(g)

    away_game_sorted.sort(key=lambda x: x.time, reverse=True)
    away_game_sorted = away_game_sorted[:5]

    for match in away_game_sorted:
        away_score = 0
        other_score = 0
        partite.append({"data": match.time, 'squadre': match.home + '-' + match.away,
                        'risultato': match.result})
        if match.home == away_team:
            away_score = str(match.result).split("-")[0]
            other_score = str(match.result).split("-")[1]
        elif match.away == away_team:
            away_score = str(match.result).split("-")[1]
            other_score = str(match.result).split("-")[0]

        # partite.append({"data": match['time'], 'squadre': match['home']['name'] + '-' + match['away']['name'],
        #                 'risultato': match['ss']})
        # if match['home']['name'] == away_team:
        #     away_score = str(match['ss']).split("-")[0]
        #     other_score = str(match['ss']).split("-")[1]
        # elif match['away']['name'] == away_team:
        #     away_score = str(match['ss']).split("-")[1]
        #     other_score = str(match['ss']).split("-")[0]

        ris = int(away_score) + int(other_score)

        # Calcolo vittorie Ospite, sconfitte Ospite e pareggi Ospite, over Ospite ed under Ospite
        if away_score > other_score:
            vittorieOspite = vittorieOspite + 1
        elif away_score < other_score:
            sconfitteOspite = sconfitteOspite + 1
        else:
            pareggiOspite = pareggiOspite + 1

        if ris > 2:
            overOspite = overOspite + 1
        else:
            underOspite = underOspite + 1


    # Inserimento if se favorita = 1
    if favorita == 1:
        if sconfitteOspite == 5:
            print('Ospite 5 sconfitte')
        if vittorieOspite == 0:
            print('Ospite non vince da 5')

    # Inserimento if se favorita = 2
    if favorita == 2:
        if vittorieOspite == 5:
            print('Ospite 5 vittorie')
        if sconfitteOspite == 0:
            print('Ospite non perde da 5')

    return partite, vittorieOspite, sconfitteOspite, underOspite, overOspite


def statisticheCasaOspite(favorita, statoFormaUnder, vittorieCasa, vittorieOspite, sconfitteCasa, sconfitteOspite,
                          underCasa, underOspite, overCasa, overOspite):
    res = ""
    if favorita == 1:
        if vittorieCasa > soglia15 and sconfitteCasa < soglia16 and sconfitteOspite > soglia17:
            # print('Segno 1X')
            res += "1x  \t"
        if vittorieCasa > soglia18 and sconfitteCasa < soglia19 and sconfitteOspite > soglia20 and overCasa > soglia21 and overOspite > soglia22:
            # print('Segno Over')
            res += "Over  \t"
        if sconfitteCasa > soglia23 and sconfitteOspite > soglia24:
            # print('Over da rosso-rosso')
            res += "Over da rosso-rosso  \t"
    if favorita == 2:
        if vittorieOspite > soglia25 and sconfitteOspite < soglia26 and sconfitteCasa > soglia27:
            # print('Segno 2')
            res += "2  \t"
        if vittorieOspite > soglia28 and sconfitteOspite < soglia29 and sconfitteCasa > soglia30 and overOspite > soglia31 and overCasa > soglia32:
            # print('Segno Over')
            res += "Over \t"
        if sconfitteCasa > soglia23 and sconfitteOspite > soglia24:
            # print('Over da rosso-rosso')
            res += "Over da rosso-rosso \t"
    if statoFormaUnder == 1:
        if vittorieCasa > soglia33 and vittorieOspite > soglia34 and underCasa > soglia35 and underOspite > soglia36:
            # print('Segno Under')
            res += "Under \t"

    return res


if __name__ == '__main__':
    risultati = openpyxl.Workbook()
    sheet = risultati.active
    sheet.title = 'Pronostici'

    # Aggiunta Header
    sheet.cell(row=1, column=1).value = 'Partita'
    sheet.cell(row=1, column=2).value = 'Data'
    sheet.cell(row=1, column=3).value = 'Ora'
    sheet.cell(row=1, column=4).value = 'Campionato'
    sheet.cell(row=1, column=5).value = 'Nazione'
    sheet.cell(row=1, column=6).value = 'Squadra Favorita'
    sheet.cell(row=1, column=7).value = 'Percentuali 1X2'
    sheet.cell(row=1, column=8).value = 'Stato Forma'
    sheet.cell(row=1, column=9).value = 'Precedenti storici'
    sheet.cell(row=1, column=10).value = 'Statistiche casa/ospite se Favorita = squadra 1'
    sheet.cell(row=1, column=11).value = 'Statistiche casa/ospite se Favorita = squadra 2'
    sheet.cell(row=1, column=12).value = 'Statistiche casa/ospite se Stato Forma = Under'

    data_da_analizzare = ""
    try:
        data_da_analizzare = sys.argv[1]
        print(data_da_analizzare)
    except Exception as e:
        print(e)

    date_format = "%Y%m%d"

    # Convert string to datetime using strptime
    date_obj = datetime.datetime.strptime(data_da_analizzare, date_format)
    print(date_obj)

    if date_obj < now:
        json_of_the_day = getEndedDayEvents(str(data_da_analizzare))
    else:
        json_of_the_day = getDayEvents(str(data_da_analizzare))

    #TODO: ATTENZIONE AL FORMATO! DEVE ESSERE ANNOMESEGIORNO ES: 20230819
    #json_of_the_day = getDayEvents(str(data_da_analizzare))

    x = 0
    for match in json_of_the_day['results']:
        home_team = match['home']['name']
        home_team_id = match['home']['id']
        away_team = match['away']['name']
        away_team_id = match['away']['id']
        match_teams = str(home_team) + " - " + str(away_team)
        match_date = datetime.datetime.fromtimestamp(int(match['time'])).strftime('%d-%m-%y')
        match_hour = datetime.datetime.fromtimestamp(int(match['time'])).strftime('%H:%M')
        match_league = match['league']['name']
        match_league_id = match['league']['id']
        match_nation = match['league']['cc']
        match_current_round = getNumberOfDayPlayed(match_league_id)

        home_points, home_pos = getPointTeam(match_league_id, home_team_id)
        away_points, away_pos = getPointTeam(match_league_id, away_team_id)

        print(match['id'], match_teams, match_date, match_hour, match_league, match_nation, match_league_id,
              "current_round:", match_current_round, "home_pos:", home_pos, "home_point:", home_points, "away_pos:", away_pos, "away_points:", away_points)

         #if match_current_round >= 5:

        getOddMatch(match['id'])
        #favorita, favoritaNome, statoFormaUnder, res_string, precedenti_string, vittorieCasa, sconfitteCasa, overCasa, underCasa, vittorieOspite, sconfitteOspite, underOspite, vittorieOspite, overOspite, numero_precedenti
        favorita, favoritaNome, statoFormaUnder, res_string, precedenti_string, vittorieCasa, sconfitteCasa, overCasa, underCasa, vittorieOspite, sconfitteOspite, underOspite, overOspite, numero_precedenti = getHistoryForDayEvent(
            match['id'], home_team, away_team)

        sheet.append(
            [match_teams, match_date, match_hour, match_league, match_nation, favoritaNome, getOddMatch(match['id']),
             res_string, precedenti_string,
             statisticheCasaOspite(favorita, statoFormaUnder, vittorieCasa, vittorieOspite, sconfitteCasa,
                                   sconfitteOspite, underCasa, underOspite, overCasa, overOspite),
             statisticheCasaOspite(favorita, statoFormaUnder, vittorieCasa, vittorieOspite, sconfitteCasa,
                                   sconfitteOspite, underCasa, underOspite, overCasa, overOspite),
             statisticheCasaOspite(favorita, statoFormaUnder, vittorieCasa, vittorieOspite, sconfitteCasa,
                                   sconfitteOspite, underCasa, underOspite, overCasa, overOspite)])

        #    print("############## MENO DI 5 PARITE: #"+str(getNumberOfDayPlayed(match_league_id)))

        risultati.save('Risultati.xlsx')


